"""Populate sample grades across all 10 trade sheets."""
import openpyxl

wb = openpyxl.load_workbook('keywords_and_phrases_checklist.xlsx')

for sn in wb.sheetnames:
    if sn in ('Summary', 'Instructions'):
        continue
    ws = wb[sn]
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 3, 'Red' if (r - 2) % 2 == 0 else 'Yellow')

wb.save('keywords_and_phrases_checklist.xlsx')

# Verify
wb2 = openpyxl.load_workbook('keywords_and_phrases_checklist.xlsx', data_only=True)
tr, ty = 0, 0
for sn in wb2.sheetnames:
    if sn in ('Summary', 'Instructions'):
        continue
    ws = wb2[sn]
    r = y = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        s = str(row[2]).strip().lower() if row[2] else ''
        if s == 'red':
            r += 1
        elif s == 'yellow':
            y += 1
    tr += r
    ty += y
    print(f'{sn}: {r} Red, {y} Yellow')
print(f'Total: {tr} Red, {ty} Yellow, {tr+ty} graded')
wb2.close()
