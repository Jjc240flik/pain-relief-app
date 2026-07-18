"""Generate keywords_and_phrases_checklist.xlsx with 10 trade sheets + summary."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── Styling ──
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
EVEN_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ODD_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
TITLE_FONT = Font(name='Calibri', bold=True, size=14, color='2F5496')
SECTION_FONT = Font(name='Calibri', bold=True, size=12, color='2F5496')
BODY_FONT = Font(name='Calibri', size=11)
THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)

# ── Trade data ──
TRADES = [
    ("Foundation Concrete", "Foundation / Concrete"),
    ("Framing", "Framing"),
    ("Plumbing Rough", "Plumbing Rough"),
    ("HVAC Rough", "HVAC Rough"),
    ("Electrical Rough", "Electrical Rough"),
    ("Drywall Plaster", "Drywall / Plaster"),
    ("Paint", "Paint"),
    ("Flooring", "Flooring"),
    ("Cabinets", "Cabinets"),
    ("Finish Work", "Finish Work"),
]

# All 400 entries: (term, type, note) for each trade
DATA = {
    "Foundation Concrete": [
        # Keywords
        ("rebar", "Word", "Missing or wrong rebar placement"),
        ("form", "Word", "Form boards shifting or breaking"),
        ("pour", "Word", "Concrete pour issues, scheduling"),
        ("cure", "Word", "Concrete not curing properly"),
        ("slab", "Word", "Slab thickness, cracks, finish"),
        ("footer", "Word", "Footer depth, width, or location"),
        ("vapor", "Word", "Vapor barrier missing or torn"),
        ("gravel", "Word", "Gravel base not compacted"),
        ("grade", "Word", "Grading issues around foundation"),
        ("mud", "Word", "Muddy conditions affecting pour"),
        ("frost", "Word", "Frost in ground, frost heave"),
        ("settle", "Word", "Settlement cracks or movement"),
        ("crack", "Word", "Any crack in foundation"),
        ("cold", "Word", "Cold weather pour complications"),
        ("honeycomb", "Word", "Honeycombing in concrete"),
        ("bullfloat", "Word", "Finish issues, surface defects"),
        ("anchor", "Word", "Anchor bolts out of position"),
        ("waterproof", "Word", "Waterproofing issues"),
        ("drain", "Word", "Drain tile or weeping tile problems"),
        ("pump", "Word", "Concrete pump truck issues"),
        # Phrases
        ("forms blew out", "Phrase", "Form failure during pour"),
        ("concrete is settling", "Phrase", "Settlement issues after pour"),
        ("slab is cracking", "Phrase", "Cracks in the slab"),
        ("rebar is wrong size", "Phrase", "Wrong rebar delivered or placed"),
        ("foundation is sinking", "Phrase", "Major structural concern"),
        ("footer not deep enough", "Phrase", "Footer below frost line issue"),
        ("cold joint in pour", "Phrase", "Cold joint from delayed pour"),
        ("honeycomb in wall", "Phrase", "Honeycombing in foundation wall"),
        ("anchor bolts off", "Phrase", "Bolts not where they need to be"),
        ("water in the hole", "Phrase", "Water accumulating before pour"),
        ("ground is frozen", "Phrase", "Frost in ground, can't dig"),
        ("mud too wet", "Phrase", "Mix too wet, slump issue"),
        ("concrete truck delayed", "Phrase", "Truck running late"),
        ("pump truck broke down", "Phrase", "Pump issues on pour day"),
        ("vapor barrier ripped", "Phrase", "Vapor barrier damaged"),
        ("grade is wrong", "Phrase", "Grading not to spec"),
        ("can't pour today", "Phrase", "Weather or ground conditions"),
        ("forms not stripped yet", "Phrase", "Forms still up, delaying next step"),
        ("foundation needs repair", "Phrase", "General foundation defect"),
        ("concrete not curing", "Phrase", "Curing time or temperature issue"),
    ],
    "Framing": [
        ("truss", "Word", "Damaged, missing, or wrong trusses"),
        ("rafter", "Word", "Rafter issues, spacing, damage"),
        ("joist", "Word", "Floor or ceiling joist problems"),
        ("stud", "Word", "Stud spacing, damage, warping"),
        ("beam", "Word", "Beam placement, size, cracks"),
        ("header", "Word", "Window/door header wrong size"),
        ("sheathing", "Word", "OSB or plywood issues"),
        ("nailing", "Word", "Nail pattern, missing nails"),
        ("hanger", "Word", "Joist hangers missing or wrong"),
        ("ridge", "Word", "Ridge beam alignment"),
        ("wall", "Word", "Wall placement, square, plumb"),
        ("corner", "Word", "Corner framing, backing"),
        ("blocking", "Word", "Missing fire blocking"),
        ("web", "Word", "Truss web damage"),
        ("notch", "Word", "Notched studs or joists"),
        ("bore", "Word", "Oversized holes in framing"),
        ("rack", "Word", "Wall racking, out of square"),
        ("plumb", "Word", "Walls or posts not plumb"),
        ("shear", "Word", "Shear wall nailing or panel issues"),
        ("holdown", "Word", "Holdown or strap missing"),
        # Phrases
        ("truss is cracked", "Phrase", "Cracked truss member"),
        ("truss is broken", "Phrase", "Broken truss during install"),
        ("wall is out of square", "Phrase", "Wall not square"),
        ("ridge beam is off", "Phrase", "Ridge not aligned"),
        ("roof sheathing damaged", "Phrase", "OSB or plywood delaminating"),
        ("joist hangers missing", "Phrase", "Hangers not installed"),
        ("beam not big enough", "Phrase", "Beam size wrong for span"),
        ("bearing point missing", "Phrase", "Load not transferred properly"),
        ("wall is not plumb", "Phrase", "Wall leaning"),
        ("nailing pattern wrong", "Phrase", "Too few nails, wrong spacing"),
        ("truss web cracked", "Phrase", "Web member failure"),
        ("header too short", "Phrase", "Header doesn't span opening"),
        ("fire blocking missing", "Phrase", "Blocking not installed"),
        ("can't set trusses today", "Phrase", "Weather or crane issues"),
        ("roof is sagging", "Phrase", "Ridge sag, structural concern"),
        ("framing failed inspection", "Phrase", "Failed city/county inspection"),
        ("shear wall nailing wrong", "Phrase", "Shear panel fasteners incorrect"),
        ("holdown not installed", "Phrase", "Seismic tie-down missing"),
        ("wall is leaning", "Phrase", "Wall out of plumb"),
        ("splice not heavy enough", "Phrase", "Splice plate undersized"),
    ],
    "Plumbing Rough": [
        ("vent", "Word", "Vent pipe routing or size issues"),
        ("drain", "Word", "Drain line slope or connection"),
        ("pipe", "Word", "Pipe damage, wrong size"),
        ("trap", "Word", "P-trap clearance or location"),
        ("stub", "Word", "Stub out location for fixtures"),
        ("meter", "Word", "Water meter location"),
        ("sewer", "Word", "Sewer connection or cleanout"),
        ("gas", "Word", "Gas piping issues"),
        ("waste", "Word", "Waste line problems"),
        ("rough", "Word", "Rough-in height or location"),
        ("tub", "Word", "Tub valve or drain location"),
        ("shower", "Word", "Shower pan or drain"),
        ("toilet", "Word", "Toilet flange or rough spacing"),
        ("sink", "Word", "Sink drain or supply"),
        ("cleanout", "Word", "Missing or inaccessible cleanout"),
        ("backwater", "Word", "Backwater valve missing"),
        ("pressure", "Word", "Water pressure issues"),
        ("valve", "Word", "Shutoff valve location or missing"),
        ("leak", "Word", "Leak at any plumbing connection"),
        ("flange", "Word", "Toilet flange height or damage"),
        # Phrases
        ("vent pipe in the way", "Phrase", "Vent conflicts with framing"),
        ("drain line has low spots", "Phrase", "Drain not sloped properly"),
        ("stub out in wrong spot", "Phrase", "Fixture stub in wrong location"),
        ("tub valve is too low", "Phrase", "Valve height off"),
        ("shower drain off center", "Phrase", "Drain not centered in shower pan"),
        ("toilet flange too high", "Phrase", "Flange height above finish floor"),
        ("cleanout is behind wall", "Phrase", "Cleanout not accessible"),
        ("gas line not capped", "Phrase", "Open gas line, safety issue"),
        ("sewer cleanout missing", "Phrase", "No cleanout at foundation exit"),
        ("rough inspection failed", "Phrase", "Failed plumbing rough inspection"),
        ("trap arm too long", "Phrase", "Trap arm exceeds max length"),
        ("backwater valve needed", "Phrase", "Required but not installed"),
        ("pipe is leaking", "Phrase", "Leak at joint or fitting"),
        ("no hot water line", "Phrase", "Missing hot water stub"),
        ("PRV not set right", "Phrase", "Pressure too high or low"),
        ("shower valve backwards", "Phrase", "Valve orientation wrong"),
        ("waste line no slope", "Phrase", "Flat waste line, will clog"),
        ("water line too small", "Phrase", "Supply line undersized"),
        ("can't rough in yet", "Phrase", "Framing not ready for plumbing"),
        ("floor drain missing", "Phrase", "Required floor drain not installed"),
    ],
    "HVAC Rough": [
        ("duct", "Word", "Ductwork size, routing, damage"),
        ("vent", "Word", "Supply or return vent location"),
        ("return", "Word", "Return air duct sizing"),
        ("plenum", "Word", "Plenum box issues"),
        ("register", "Word", "Register boot location"),
        ("trunk", "Word", "Main trunk line issues"),
        ("branch", "Word", "Branch duct tap issues"),
        ("flex", "Word", "Flex duct kinked or crushed"),
        ("chase", "Word", "Chase not framed correctly"),
        ("condenser", "Word", "Condenser pad or location"),
        ("furnace", "Word", "Furnace placement or clearance"),
        ("flue", "Word", "Flue pipe clearance issues"),
        ("thermostat", "Word", "Thermostat wire location"),
        ("zone", "Word", "Zoning damper issues"),
        ("filter", "Word", "Filter size or access issues"),
        ("line set", "Word", "Refrigerant line set routing"),
        ("coil", "Word", "Evaporator coil issues"),
        ("heat", "Word", "Heat pump or furnace not heating"),
        ("cool", "Word", "AC not cooling"),
        ("condensate", "Word", "Condensate drain line issues"),
        # Phrases
        ("duct is crushed", "Phrase", "Duct crushed during framing"),
        ("flex duct is kinked", "Phrase", "Kinked flex, airflow restricted"),
        ("return duct too small", "Phrase", "Return undersized for unit"),
        ("register boot in wrong spot", "Phrase", "Boot location conflicts with cabinet"),
        ("trunk line leaking", "Phrase", "Duct leakage at seams"),
        ("chase not big enough", "Phrase", "Chase too small for duct"),
        ("fresh air intake blocked", "Phrase", "Intake covered or obstructed"),
        ("condenser pad not level", "Phrase", "Pad settling, unit not level"),
        ("flue clearance too close", "Phrase", "Flue too close to combustibles"),
        ("no thermostat wire", "Phrase", "Wire not run to location"),
        ("zone damper not working", "Phrase", "Damper stuck or wired wrong"),
        ("duct inspection failed", "Phrase", "Duct leakage test failure"),
        ("furnace closet too small", "Phrase", "Clearance not to code"),
        ("line set has a kink", "Phrase", "Refrigerant line damaged"),
        ("no power at unit", "Phrase", "Electric not run yet"),
        ("filter access blocked", "Phrase", "Can't change filter after build"),
        ("supply duct blowing off", "Phrase", "Duct disconnected at boot"),
        ("condensate line clogged", "Phrase", "Drain line blocked, water risk"),
        ("unit not cooling", "Phrase", "AC not working"),
        ("can't rough in today", "Phrase", "Framing not ready for HVAC"),
    ],
    "Electrical Rough": [
        ("wire", "Word", "Wire gauge, type, or damage"),
        ("Romex", "Word", "Romex stapling, damage violations"),
        ("box", "Word", "Box fill, placement, missing"),
        ("switch", "Word", "Switch location or height"),
        ("outlet", "Word", "Outlet spacing or location"),
        ("receptacle", "Word", "Receptacle type or placement"),
        ("conduit", "Word", "Conduit fill, routing, damage"),
        ("panel", "Word", "Panel location, clearance"),
        ("breaker", "Word", "Breaker sizing or trip issues"),
        ("ground", "Word", "Missing or improper ground"),
        ("AFCI", "Word", "AFCI breaker requirements"),
        ("GFCI", "Word", "GFCI protection missing"),
        ("smoke", "Word", "Smoke detector location, wiring"),
        ("staple", "Word", "Stapling too tight, spacing"),
        ("plate", "Word", "Missing nail plates on studs"),
        ("drill", "Word", "Oversized holes in joists/studs"),
        ("junction", "Word", "Junction box accessibility"),
        ("meter", "Word", "Meter can location or damage"),
        ("home run", "Word", "Missing home run to panel"),
        ("splice", "Word", "Improper splice in junction box"),
        # Phrases
        ("wire is too small", "Phrase", "Undersized for circuit"),
        ("outlet box in wrong spot", "Phrase", "Box doesn't line up with layout"),
        ("switch height is wrong", "Phrase", "Not ADA or spec height"),
        ("no ground wire", "Phrase", "Ungrounded circuit"),
        ("nail plate missing", "Phrase", "Wire at risk of nail puncture"),
        ("holes drilled too big", "Phrase", "Oversized holes in framing"),
        ("Romex not stapled", "Phrase", "Loose wire, code violation"),
        ("smoke detector not tied", "Phrase", "Not interconnected"),
        ("panel clearance too tight", "Phrase", "Panel space not to code"),
        ("GFCI protection missing", "Phrase", "Required but not installed"),
        ("can't rough in yet", "Phrase", "Framing/walls not ready"),
        ("rough inspection failed", "Phrase", "Failed electrical inspection"),
        ("junction box buried", "Phrase", "J-box not accessible"),
        ("conduit is full", "Phrase", "Conduit fill over max"),
        ("breaker keeps tripping", "Phrase", "Something is shorting"),
        ("no power at panel", "Phrase", "Utility not connected"),
        ("meter can is damaged", "Phrase", "Meter socket damaged"),
        ("need a home run", "Phrase", "Missing home run to panel"),
        ("arc fault missing", "Phrase", "AFCI not installed"),
        ("light box not supported", "Phrase", "Box not rated for ceiling fan"),
    ],
    "Drywall Plaster": [
        ("mud", "Word", "Joint compound, drying, finish"),
        ("tape", "Word", "Tape bubbling, lifting, cracking"),
        ("bead", "Word", "Corner bead loose or damaged"),
        ("screw", "Word", "Screw pops, missed, or depth"),
        ("nail", "Word", "Nail pops"),
        ("joint", "Word", "Joint finishing, cracking"),
        ("corner", "Word", "Corner issues, inside/outside"),
        ("ceiling", "Word", "Ceiling sag, seam issues"),
        ("texture", "Word", "Texture matching, orange peel"),
        ("sand", "Word", "Sanding dust, finish quality"),
        ("patch", "Word", "Patch work needed"),
        ("crack", "Word", "Cracks in drywall"),
        ("bump", "Word", "Bump at joint or repair"),
        ("bow", "Word", "Wall or ceiling bowing"),
        ("hump", "Word", "Hump at tapered joint"),
        ("base", "Word", "Gap at base of wall"),
        ("rust", "Word", "Rust spots from exposed nails"),
        ("pop", "Word", "Screw or nail pop"),
        ("float", "Word", "Floating or skim coat issues"),
        ("fire", "Word", "Fire-rated drywall or tape issues"),
        # Phrases
        ("tape is bubbling", "Phrase", "Tape not bedded properly"),
        ("corner bead is loose", "Phrase", "Bead not nailed properly"),
        ("screw pops everywhere", "Phrase", "Too many exposed screw heads"),
        ("ceiling is sagging", "Phrase", "Ceiling board not supported"),
        ("joints are cracking", "Phrase", "Tapered joint cracks"),
        ("texture doesn't match", "Phrase", "Patch texture different"),
        ("mud isn't dry yet", "Phrase", "Can't sand or paint"),
        ("truss uplift cracks", "Phrase", "Cracks at ceiling joints"),
        ("need to skim the wall", "Phrase", "Wall needs skim coat"),
        ("corner has a bump", "Phrase", "Corner finish not smooth"),
        ("base gap is too big", "Phrase", "Gap at bottom of wall"),
        ("nail rust coming through", "Phrase", "Rust spots bleeding through"),
        ("drywall is wet", "Phrase", "Water damage to installed board"),
        ("can't start hanging yet", "Phrase", "Rough trades not finished"),
        ("ceiling board has a bow", "Phrase", "Board not straight"),
        ("fire tape incomplete", "Phrase", "Fire-rated wall tape not done"),
        ("sanding dust everywhere", "Phrase", "Dust covering job site"),
        ("drywall damaged in corner", "Phrase", "Crushed or broken corner"),
        ("need a re-tape", "Phrase", "Tape joint failed, needs redo"),
        ("screw is stripped", "Phrase", "Screw not catching in stud"),
    ],
    "Paint": [
        ("wall", "Word", "Wall paint coverage or issues"),
        ("ceiling", "Word", "Ceiling paint, flashing"),
        ("trim", "Word", "Trim paint, brush marks, drips"),
        ("door", "Word", "Door paint, finish quality"),
        ("cabinet", "Word", "Cabinet paint or color issues"),
        ("stain", "Word", "Stain color or application"),
        ("primer", "Word", "Primer not applied or wrong type"),
        ("coat", "Word", "Number of coats, coverage"),
        ("flash", "Word", "Shiny spots from touch-up"),
        ("drip", "Word", "Paint drips on trim or floor"),
        ("roller", "Word", "Roller marks or texture"),
        ("brush", "Word", "Brush marks visible"),
        ("gap", "Word", "Caulk gap at trim"),
        ("caulk", "Word", "Caulking cracking or missing"),
        ("sheen", "Word", "Sheen mismatch"),
        ("patch", "Word", "Patched area not primed"),
        ("overspray", "Word", "Paint spray reaching adjacent areas"),
        ("peel", "Word", "Paint peeling or adhesion failure"),
        ("bleed", "Word", "Paint bleeding under tape"),
        ("touch up", "Word", "Final touch-up not done"),
        # Phrases
        ("wall needs another coat", "Phrase", "Coverage not sufficient"),
        ("ceiling is flashing", "Phrase", "Shiny spots on ceiling"),
        ("trim has brush marks", "Phrase", "Visible brush strokes"),
        ("paint is peeling", "Phrase", "Adhesion failure"),
        ("caulk is cracking", "Phrase", "Caulk failed at joints"),
        ("color doesn't match", "Phrase", "Touch-up paint different shade"),
        ("sheen is wrong", "Phrase", "Eggshell vs flat mismatch"),
        ("overspray on the floor", "Phrase", "Paint droplets on flooring"),
        ("cabinet doors have drips", "Phrase", "Runs or sags on doors"),
        ("stain came out blotchy", "Phrase", "Stain uneven absorption"),
        ("patch wasn't primed", "Phrase", "Patch flashing through paint"),
        ("paint bleeding under tape", "Phrase", "Tape line not crisp"),
        ("wall texture doesn't match", "Phrase", "Orange peel vs smooth mismatch"),
        ("drywall showing through", "Phrase", "Joints visible after painting"),
        ("need to repaint the room", "Phrase", "Full repaint needed"),
        ("can't paint yet", "Phrase", "Drywall not ready"),
        ("exterior paint peeling", "Phrase", "Exterior paint failing"),
        ("fence not painted", "Phrase", "Missed scope item"),
        ("touch up not done", "Phrase", "Final touch-up walk not complete"),
        ("garage floor not painted", "Phrase", "Missed scope item"),
    ],
    "Flooring": [
        ("plank", "Word", "LVP or laminate plank issues"),
        ("tile", "Word", "Tile crack, lippage, layout"),
        ("hardwood", "Word", "Hardwood cupping, gaps"),
        ("carpet", "Word", "Carpet seams, stretching"),
        ("vinyl", "Word", "Vinyl sheet or plank issues"),
        ("grout", "Word", "Grout cracking, color issues"),
        ("subfloor", "Word", "Subfloor prep, flatness"),
        ("underlayment", "Word", "Underlayment type or gaps"),
        ("transition", "Word", "Transition strip issues"),
        ("baseboard", "Word", "Baseboard gaps or damage"),
        ("lippage", "Word", "Tile height difference"),
        ("seam", "Word", "Carpet or vinyl seam visible"),
        ("gap", "Word", "Gap between planks or at wall"),
        ("buckle", "Word", "Floor buckling or tenting"),
        ("squeak", "Word", "Floor squeaking after install"),
        ("level", "Word", "Floor not level or flat"),
        ("moisture", "Word", "Moisture in subfloor or slab"),
        ("expansion", "Word", "Expansion gap missing"),
        ("glue", "Word", "Glue-down issues, adhesive failure"),
        ("padding", "Word", "Carpet padding issues"),
        # Phrases
        ("planks are separating", "Phrase", "Gaps between LVP planks"),
        ("tile has lippage", "Phrase", "Height difference at tile edges"),
        ("grout is cracking", "Phrase", "Grout failing in traffic areas"),
        ("hardwood is cupping", "Phrase", "Wood absorbing moisture"),
        ("carpet seam is visible", "Phrase", "Seam shows through"),
        ("subfloor is not flat", "Phrase", "Requires self-leveler"),
        ("transition strip missing", "Phrase", "Missing at doorways"),
        ("floor is squeaking", "Phrase", "Squeaks after install"),
        ("expansion gap missing", "Phrase", "No gap at perimeter"),
        ("baseboard has gaps", "Phrase", "Gaps at floor or wall"),
        ("underlayment is wrinkled", "Phrase", "Wrinkles show through flooring"),
        ("moisture in the slab", "Phrase", "Can't install until dry"),
        ("tile is cracked", "Phrase", "Cracked tile, needs replacement"),
        ("pattern is off", "Phrase", "Tile pattern not aligned"),
        ("can't install yet", "Phrase", "Paint or other trades not done"),
        ("floor is buckling", "Phrase", "Tent or buckle at seams"),
        ("glue is not holding", "Phrase", "Adhesive failure"),
        ("leveling compound cracked", "Phrase", "Self-leveler failed"),
        ("flooring not in stock", "Phrase", "Material delay from supplier"),
        ("carpet padding tore", "Phrase", "Padding damaged during install"),
    ],
    "Cabinets": [
        ("cabinet", "Word", "General cabinet issue"),
        ("door", "Word", "Door alignment, swing, damage"),
        ("drawer", "Word", "Drawer glide, fit, front"),
        ("hinge", "Word", "Hinge adjustment, damage"),
        ("pull", "Word", "Handle/pull placement or missing"),
        ("knob", "Word", "Knob location or wrong style"),
        ("shelf", "Word", "Shelf adjustment, missing pins"),
        ("face", "Word", "Cabinet face frame alignment"),
        ("panel", "Word", "End panel, filler, gable"),
        ("filler", "Word", "Filler strip size or gap"),
        ("toe", "Word", "Toe kick height or missing"),
        ("crown", "Word", "Crown molding alignment or gap"),
        ("scribe", "Word", "Scribe piece or scribing needed"),
        ("scratch", "Word", "Scratched finish"),
        ("dent", "Word", "Dent in cabinet body or door"),
        ("level", "Word", "Cabinet not level"),
        ("shim", "Word", "Shimming visible or missing"),
        ("warranty", "Word", "Warranty replacement part needed"),
        ("damage", "Word", "Damage during delivery or install"),
        ("reveal", "Word", "Door reveal uneven"),
        # Phrases
        ("door is not aligned", "Phrase", "Door hitting adjacent door"),
        ("drawer won't close", "Phrase", "Drawer glide issue or obstruction"),
        ("hinge is broken", "Phrase", "Hinge damaged during install"),
        ("crown molding has gap", "Phrase", "Gap at ceiling or cabinet top"),
        ("toe kick not right", "Phrase", "Toe kick height or depth off"),
        ("filler strip missing", "Phrase", "Gap between cabinet and wall"),
        ("cabinet not level", "Phrase", "Cabinet tilted or rocking"),
        ("pull placement wrong", "Phrase", "Pulls not centered on doors"),
        ("shelf pins missing", "Phrase", "Adjustable shelves not installed"),
        ("end panel scratched", "Phrase", "Visible scratch on exposed panel"),
        ("drawer glides noisy", "Phrase", "Soft-close not working"),
        ("cabinet door warped", "Phrase", "Door bowed or twisted"),
        ("scribe piece needed", "Phrase", "Gap at wall, needs scribing"),
        ("counter not templated", "Phrase", "Template for counters not done"),
        ("backsplash not installed", "Phrase", "Backsplash gap above counter"),
        ("island cabinet off center", "Phrase", "Island position not aligned"),
        ("can't install yet", "Phrase", "Floor or walls not ready"),
        ("cabinet damaged in delivery", "Phrase", "Shipping damage"),
        ("door swing conflicts", "Phrase", "Door hits wall or appliance"),
        ("opening not sized right", "Phrase", "Opening too small for appliance"),
    ],
    "Finish Work": [
        ("trim", "Word", "Base, casing, or crown issues"),
        ("casing", "Word", "Window or door casing gaps"),
        ("base", "Word", "Baseboard gaps or miters"),
        ("crown", "Word", "Crown molding miters or copes"),
        ("miter", "Word", "Miter joint gap or angle"),
        ("cope", "Word", "Coped joint quality"),
        ("caulk", "Word", "Caulking needed or cracking"),
        ("nail", "Word", "Nail holes, pops, or fill"),
        ("putty", "Word", "Putty not applied or color wrong"),
        ("stain", "Word", "Stain match, blotchiness"),
        ("clear", "Word", "Clear coat or polyurethane issues"),
        ("stair", "Word", "Stair tread, riser, or rail"),
        ("railing", "Word", "Railing height or loose"),
        ("baluster", "Word", "Baluster spacing or loose"),
        ("handrail", "Word", "Handrail height or brackets"),
        ("column", "Word", "Column wraps or damage"),
        ("mantel", "Word", "Fireplace mantel install"),
        ("shelf", "Word", "Built-in shelf or niche"),
        ("hardware", "Word", "Door hardware, strike plates"),
        ("threshold", "Word", "Door threshold height or gap"),
        # Phrases
        ("miter joint is open", "Phrase", "Gap at corner joint"),
        ("casing doesn't sit flush", "Phrase", "Casing gap at jamb"),
        ("baseboard has gap at floor", "Phrase", "Gap under baseboard"),
        ("crown molding miter off", "Phrase", "Crown corner not tight"),
        ("nail holes not filled", "Phrase", "Putty not applied yet"),
        ("caulk is cracking", "Phrase", "Caulk at trim-to-wall failed"),
        ("stain color doesn't match", "Phrase", "Touch-up stain different shade"),
        ("handrail is loose", "Phrase", "Railing not secured properly"),
        ("baluster spacing too wide", "Phrase", "Exceeds code spacing"),
        ("stair tread is squeaking", "Phrase", "Tread not glued/nailed"),
        ("door doesn't latch", "Phrase", "Strike plate not aligned"),
        ("door is dragging", "Phrase", "Door rubbing at bottom or side"),
        ("weatherstripping missing", "Phrase", "Door seal not installed"),
        ("column wrap damaged", "Phrase", "Dent or scratch on column"),
        ("mantel not centered", "Phrase", "Fireplace surround off"),
        ("built-in shelf not level", "Phrase", "Shelf sagging or crooked"),
        ("hardware finish wrong", "Phrase", "Wrong color/style installed"),
        ("can't start finish yet", "Phrase", "Paint or other trades not done"),
        ("touch up paint needed", "Phrase", "Final touch-up after install"),
        ("window sill not caulked", "Phrase", "Sill gap, not sealed"),
    ],
}


# ============================================================
# SHEET 1: Summary
# ============================================================
ws_summary = wb.active
ws_summary.title = "Summary"
ws_summary.sheet_properties.tabColor = "2F5496"

# Title
ws_summary.merge_cells('A1:E1')
ws_summary['A1'] = 'Keywords & Phrases — Severity Grading Summary'
ws_summary['A1'].font = TITLE_FONT
ws_summary['A1'].alignment = Alignment(horizontal='center')
ws_summary.row_dimensions[1].height = 30

# Subtitle
ws_summary.merge_cells('A2:E2')
ws_summary['A2'] = 'Select Yellow or Red on each trade sheet. This summary updates automatically.'
ws_summary['A2'].font = Font(name='Calibri', italic=True, size=10, color='666666')
ws_summary['A2'].alignment = Alignment(horizontal='center')

# Headers (row 4)
headers_s = ['Trade', 'Total Items', 'Yellow Count', 'Red Count', '% Graded']
for col, h in enumerate(headers_s, 1):
    cell = ws_summary.cell(row=4, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')
    cell.border = THIN_BORDER
ws_summary.row_dimensions[4].height = 22

# Column widths
ws_summary.column_dimensions['A'].width = 22
ws_summary.column_dimensions['B'].width = 14
ws_summary.column_dimensions['C'].width = 16
ws_summary.column_dimensions['D'].width = 14
ws_summary.column_dimensions['E'].width = 14

# Sheet names — must match the tab names we'll use
SHEET_NAMES = [
    "Foundation Concrete",
    "Framing",
    "Plumbing Rough",
    "HVAC Rough",
    "Electrical Rough",
    "Drywall Plaster",
    "Paint",
    "Flooring",
    "Cabinets",
    "Finish Work",
]

# Data rows with COUNTIF formulas
for i, trade in enumerate(SHEET_NAMES):
    row = 5 + i
    total = len(DATA[trade])
    ws_summary.cell(row=row, column=1, value=trade).font = BODY_FONT
    ws_summary.cell(row=row, column=1).border = THIN_BORDER
    ws_summary.cell(row=row, column=2, value=total).font = BODY_FONT
    ws_summary.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=2).border = THIN_BORDER

    # COUNTIF formulas — count "Yellow" and "Red" in column C of each trade sheet
    last_data_row = total + 1  # header is row 1, data starts row 2
    col_c = get_column_letter(3)  # Severity column = C

    ws_summary.cell(row=row, column=3).value = (
        f"=COUNTIF('{trade}'!$C$2:$C${last_data_row},\"Yellow\")"
    )
    ws_summary.cell(row=row, column=3).font = BODY_FONT
    ws_summary.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=3).border = THIN_BORDER

    ws_summary.cell(row=row, column=4).value = (
        f"=COUNTIF('{trade}'!$C$2:$C${last_data_row},\"Red\")"
    )
    ws_summary.cell(row=row, column=4).font = BODY_FONT
    ws_summary.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=4).border = THIN_BORDER

    # % Graded = (Yellow + Red) / Total
    ws_summary.cell(row=row, column=5).value = (
        f"=IF(B{row}=0,0,ROUND((C{row}+D{row})/B{row}*100,0))"
    )
    ws_summary.cell(row=row, column=5).font = BODY_FONT
    ws_summary.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=5).border = THIN_BORDER
    ws_summary.cell(row=row, column=5).number_format = '0"%"'

    # Alternating row colors
    if i % 2 == 0:
        for col in range(1, 6):
            ws_summary.cell(row=row, column=col).fill = EVEN_FILL

# Totals row
total_row = 5 + len(SHEET_NAMES)
ws_summary.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Calibri', bold=True, size=11)
ws_summary.cell(row=total_row, column=1).border = THIN_BORDER
ws_summary.cell(row=total_row, column=2, value=sum(len(DATA[t]) for t in SHEET_NAMES))
ws_summary.cell(row=total_row, column=2).font = Font(name='Calibri', bold=True, size=11)
ws_summary.cell(row=total_row, column=2).alignment = Alignment(horizontal='center')
ws_summary.cell(row=total_row, column=2).border = THIN_BORDER
for col in range(3, 6):
    ws_summary.cell(row=total_row, column=col).value = (
        f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{total_row-1})"
    )
    ws_summary.cell(row=total_row, column=col).font = Font(name='Calibri', bold=True, size=11)
    ws_summary.cell(row=total_row, column=col).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=total_row, column=col).border = THIN_BORDER
# Override % with average
ws_summary.cell(row=total_row, column=5).value = (
    f"=IF(B{total_row}=0,0,ROUND((C{total_row}+D{total_row})/B{total_row}*100,0))"
)
ws_summary.cell(row=total_row, column=5).number_format = '0"%"'

# Freeze panes (header row)
ws_summary.freeze_panes = 'A5'


# ============================================================
# SHEET 2: Instructions
# ============================================================
ws_instructions = wb.create_sheet("Instructions")
ws_instructions.sheet_properties.tabColor = "4472C4"
ws_instructions.column_dimensions['A'].width = 90

instructions_text = [
    ("Keywords and Phrases for ClassifierEngine", TITLE_FONT),
    ("", BODY_FONT),
    ("Purpose:", SECTION_FONT),
    ("This document contains the initial set of keywords and short phrases used by the ClassifierEngine to determine Red and Yellow issues. It is designed specifically for USA residential home building projects.", BODY_FONT),
    ("", BODY_FONT),
    ("Phase 1 – Foundation", SECTION_FONT),
    ("This is the first version of the keyword library. It was built using common language from USA home builders and field subcontractors.", BODY_FONT),
    ("", BODY_FONT),
    ("Future Improvements", SECTION_FONT),
    ("As the system is used, we will expand and refine this list based on:", BODY_FONT),
    ("  • Real messages coming through SMS, Email, Voice, and MMS from actual subcontractors.", BODY_FONT),
    ("  • Feedback and corrections from the Project Manager.", BODY_FONT),
    ("  • Patterns observed across multiple builders once more companies start using the system.", BODY_FONT),
    ("", BODY_FONT),
    ("This document is intended to be a living file that improves over time with real-world usage.", BODY_FONT),
    ("", BODY_FONT),
    ("━━ How to Use This File ━━", SECTION_FONT),
    ("", BODY_FONT),
    ("1. Each trade has its own sheet (Foundation Concrete, Framing, etc.).", BODY_FONT),
    ("", BODY_FONT),
    ("2. On each sheet, you will find:", BODY_FONT),
    ("     Column A → Term (the keyword or phrase)", BODY_FONT),
    ("     Column B → Type (Word or Phrase)", BODY_FONT),
    ("     Column C → Severity (dropdown: Yellow or Red)", BODY_FONT),
    ("     Column D → Notes / Example Issue", BODY_FONT),
    ("", BODY_FONT),
    ("3. For each row, click the Severity cell and choose from the dropdown:", BODY_FONT),
    ("     Yellow → Concerning but not immediately blocking", BODY_FONT),
    ("     Red → Blocks progress, safety risk, major delay, prevents next trade", BODY_FONT),
    ("", BODY_FONT),
    ("4. The Summary sheet (first tab) automatically updates as you grade:", BODY_FONT),
    ("     Total items per trade", BODY_FONT),
    ("     Yellow and Red counts (COUNTIF formulas)", BODY_FONT),
    ("     Percentage graded", BODY_FONT),
    ("", BODY_FONT),
    ("5. When finished, save the file and return it. The grading data will be", BODY_FONT),
    ("   used to update the ClassifierEngine keyword rules.", BODY_FONT),
    ("", BODY_FONT),
    ("Severity Guidance:", SECTION_FONT),
    ("  Red:   Serious issues that usually block progress, create safety risks,", BODY_FONT),
    ("         cause major delays, or prevent the next trade from starting.", BODY_FONT),
    ("  Yellow: Problems that are concerning but not immediately blocking", BODY_FONT),
    ("         (e.g. quality issues, cleanliness, minor delays).", BODY_FONT),
]

for i, (text, font) in enumerate(instructions_text):
    cell = ws_instructions.cell(row=i + 1, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True)

# Freeze
ws_instructions.freeze_panes = 'A2'


# ============================================================
# TRADE SHEETS (10)
# ============================================================
DV = DataValidation(type="list", formula1='"Yellow,Red"', allow_blank=True)
DV.error = "Please choose Yellow or Red from the dropdown."
DV.errorTitle = "Invalid Severity"
DV.prompt = "Select Yellow or Red"
DV.promptTitle = "Severity"

HEADERS = ['Term', 'Type', 'Severity', 'Notes / Example Issue']
COL_WIDTHS = [35, 12, 16, 55]

for sheet_name, display_name in TRADES:
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = "4472C4"

    # Column widths
    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Header row
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

    # Data
    entries = DATA[sheet_name]
    for i, (term, typ, note) in enumerate(entries):
        row = i + 2
        ws.cell(row=row, column=1, value=term).font = BODY_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=typ).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3).font = BODY_FONT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=note).font = BODY_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)

        # Alternating rows
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = EVEN_FILL
        else:
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = ODD_FILL

    # Data validation dropdown on Severity column (C) for all data rows
    last_row = len(entries) + 1
    dv = DataValidation(type="list", formula1='"Yellow,Red"', allow_blank=True)
    dv.error = "Please choose Yellow or Red from the dropdown."
    dv.errorTitle = "Invalid Severity"
    dv.prompt = "Select Yellow or Red"
    dv.promptTitle = "Severity"
    dv.add(f"C2:C{last_row}")
    ws.add_data_validation(dv)

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Bold count in summary row area (optional — just a visual)
    ws.cell(row=last_row + 2, column=1, value=f"Total entries: {len(entries)}").font = Font(
        name='Calibri', italic=True, size=10, color='888888'
    )


# ── Save ──
output_path = '/root/pain-relief-app/andon/keywords_and_phrases_checklist.xlsx'
wb.save(output_path)
print(f"✅ Created: {output_path}")
print(f"   Sheets: Summary, Instructions, + 10 trade sheets")
print(f"   Total entries: {sum(len(DATA[t]) for t in SHEET_NAMES)}")
print(f"   Data validation dropdowns on all trade Severity columns")
print(f"   COUNTIF formulas on Summary sheet")
