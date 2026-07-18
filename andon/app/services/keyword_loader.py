"""
Keyword Rules Loader — Reads the graded Excel checklist and exports structured rules.

Usage:
    from app.services.keyword_loader import load_keywords_from_excel
    report = load_keywords_from_excel("keywords_and_phrases_checklist.xlsx")

This produces a JSON file (keywords_rules.json) that the ClassifierEngine
can load at startup or reload on demand via /admin/import-keywords.
"""

import json
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# Expected trade sheet names in the Excel file
TRADE_SHEETS = [
    "Foundation Concrete",
    "Framing",
    "Plumbing Rough",
    "HVAC Rough",
    "Electrical Rough",
    "Drywall Plaster",
    "Paint",
    "Flooring",
    "Cabinets",
    "Finish Work",
]

RULES_FILE = Path(__file__).resolve().parent.parent / "keywords_rules.json"


def load_keywords_from_excel(excel_path: str | Path) -> dict:
    """
    Read the graded Excel checklist and extract Red/Yellow terms per trade.

    Args:
        excel_path: Path to the keywords_and_phrases_checklist.xlsx file.

    Returns:
        A dict with:
          - trades: {trade_name: {"red": [...], "yellow": [...]}}
          - summary: {trade_name: {"yellow": N, "red": N, "total": N, "graded": N}}
          - totals: {"yellow": N, "red": N, "total": N}
    """
    import openpyxl

    if not Path(excel_path).exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    trades = {}
    summary = {}
    grand_yellow = 0
    grand_red = 0
    grand_total = 0

    for sheet_name in TRADE_SHEETS:
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found in workbook — skipping.", sheet_name)
            continue

        ws = wb[sheet_name]
        red_terms = []
        yellow_terms = []
        total = 0
        graded = 0

        # Data starts at row 2 (row 1 is header)
        for row in ws.iter_rows(min_row=2, values_only=True):
            term = row[0]
            severity = row[2]  # Column C = Severity

            if not term or not str(term).strip():
                continue
            term = str(term).strip()
            total += 1
            grand_total += 1

            if not severity or str(severity).strip() == "":
                continue

            severity = str(severity).strip().lower()
            graded += 1

            if severity == "red":
                red_terms.append(term)
                grand_red += 1
            elif severity == "yellow":
                yellow_terms.append(term)
                grand_yellow += 1

        trades[sheet_name] = {
            "red": red_terms,
            "yellow": yellow_terms,
        }
        summary[sheet_name] = {
            "yellow": len(yellow_terms),
            "red": len(red_terms),
            "total": total,
            "graded": graded,
        }

        logger.info(
            "%s: %d red, %d yellow, %d/%d graded",
            sheet_name, len(red_terms), len(yellow_terms), graded, total,
        )

    wb.close()

    result = {
        "trades": trades,
        "summary": summary,
        "totals": {
            "yellow": grand_yellow,
            "red": grand_red,
            "total": grand_total,
        },
    }

    # Save to JSON file
    _save_rules(result)
    return result


def _save_rules(data: dict) -> None:
    """Persist the graded keyword rules to a JSON file."""
    rules_data = {}
    for trade_name, terms in data["trades"].items():
        rules_data[trade_name] = {
            "red_keywords": terms["red"],
            "yellow_keywords": terms["yellow"],
        }

    with open(RULES_FILE, "w") as f:
        json.dump(rules_data, f, indent=2)

    logger.info("Saved keyword rules to %s", RULES_FILE)


def load_rules_from_json() -> dict:
    """Load graded keyword rules from the persistent JSON file."""
    if not RULES_FILE.exists():
        logger.warning("No keyword rules file found at %s — using defaults.", RULES_FILE)
        return {}
    with open(RULES_FILE) as f:
        data = json.load(f)
    logger.info("Loaded %d trade rule sets from %s", len(data), RULES_FILE)
    return data


def format_report(result: dict) -> str:
    """Format a human-readable summary of the import."""
    t = result["totals"]
    lines = [
        "=" * 55,
        "📊 KEYWORD IMPORT REPORT",
        "=" * 55,
        f"Total terms processed: {t['total']}",
        f"  Yellow: {t['yellow']}",
        f"  Red:    {t['red']}",
        f"  Total graded: {t['yellow'] + t['red']} / {t['total']}",
        "",
        f"{'Trade':<28s} {'Yellow':>8s} {'Red':>8s} {'Graded':>8s}",
        "-" * 55,
    ]
    for trade, s in result["summary"].items():
        lines.append(
            f"{trade:<28s} {s['yellow']:>8d} {s['red']:>8d} {s['graded']:>8d}/{s['total']:<3d}"
        )
    lines.append("-" * 55)
    lines.append(f"Total graded: {t['yellow'] + t['red']} / {t['total']}")
    lines.append("")
    lines.append("Rules saved to: keywords_rules.json")
    lines.append("=" * 55)
    return "\n".join(lines)
